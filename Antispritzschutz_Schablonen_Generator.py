# Module einfügen

import tkinter as tk
from tkinter import END, StringVar, filedialog
import os
from PIL import Image, ImageTk
import PIL
import openpyxl
import sys
from openpyxl import Workbook
from datetime import date, datetime
import Antispritzschutz_Script_Liner_4
import Antispritzschutz_Script_Liner_5


# Erstellen vom main window
window=tk.Tk()
window_width=850
window_height=600

# Main Window in der mitte vom Bildschirm anzeigen lassen
screen_width = window.winfo_screenwidth()
screen_height = window.winfo_screenheight()
x_coord = (screen_width/2) - (window_width/2)
y_coord = (screen_height/2) - (window_height/2) -50
window.geometry("+{}+{}".format(int(x_coord), int(y_coord)))

# Main Window in 8 Spalten aufteilen um die Buttons und Eingabefelder an ein Raster anzupassen
window.configure(highlightthickness=3, highlightbackground='black')
for i in range(8):
    window.columnconfigure(i, weight=1, minsize=100)
window.geometry(str(window_width) + "x" + str(window_height))

# Arrays für die Positionen der Ausstanzungen
list_trennung_alu_NPL = []
list_rungen_NPL = []
list_trennung_schuerze_NPL = []
list_planspanner_NPL = []
list_trennung_alu_NPR = []
list_rungen_NPR = []
list_trennung_schuerze_NPR = []
list_planspanner_NPR = []
abstand_alu_schuerze = ["",""]
laenge_schuerzen = ["",""]
window_side =[False,False]

angewinkelt_rechts = tk.IntVar()     
angewinkelt_links = tk.IntVar()

list_arr_NPL= [list_trennung_alu_NPL, list_rungen_NPL, list_trennung_schuerze_NPL, list_planspanner_NPL]   
list_arr_NPR= [list_trennung_alu_NPR, list_rungen_NPR, list_trennung_schuerze_NPR, list_planspanner_NPR]   

# Funktion um die Variablen zurückzusetzen und die Widgets vom Display zu löschen
def reset_window():
    # Destroy all widgets in the window
    for widget in window.winfo_children():
        widget.destroy()
    list_trennung_alu_NPL.clear()
    list_rungen_NPL.clear()
    list_trennung_schuerze_NPL.clear()
    list_planspanner_NPL.clear()
    list_trennung_alu_NPR.clear()
    list_rungen_NPR.clear()
    list_trennung_schuerze_NPR.clear()
    list_planspanner_NPR.clear()
    abstand_alu_schuerze[:1]=["",""]
    laenge_schuerzen[:1]=["",""]
    window_side[:1] =[False,False]
    
# Funktion um die Werte aus den Eingabefeldern den Arrays hinzuzufügen
def add_string(entry,listbox_nr,list_nr):
    # Get the string from the entry widget
    string = entry.get()
    if string:
        if ((listbox_nr==0 or listbox_nr==1) and window_side[0]==False)or((listbox_nr==2 or listbox_nr==3) and window_side[1]==False):
            # Convert the string to an floating point number, sort the list and add it
            number = float(string)
            list_arr_NPL[list_nr].append(number)
            list_arr_NPL[list_nr].sort()  # Sort the list
            listbox_arr[listbox_nr].delete(0, "end")  # Clear the listbox
            for item in list_arr_NPL[list_nr]:  # Insert the sorted items into the listbox
                listbox_arr[listbox_nr].insert("end", item)
            entry.delete(0, 'end')  # Clear the entry widget

        if ((listbox_nr==0 or listbox_nr==1) and window_side[0]==True)or((listbox_nr==2 or listbox_nr==3) and window_side[1]==True):
            number = float(string)
            list_arr_NPR[list_nr].append(number)
            list_arr_NPR[list_nr].sort()  # Sort the list
            listbox_arr[listbox_nr].delete(0, "end")  # Clear the listbox
            for item in list_arr_NPR[list_nr]:  # Insert the sorted items into the listbox
                listbox_arr[listbox_nr].insert("end", item)
            entry.delete(0, 'end')  # Clear the entry widget

            
# Funktion zur Löschng der Variablen in den Arrays und den Listboxen
def delete_string(listbox_nr,list_nr):
    # Get the selected string
    selected_string = listbox_arr[listbox_nr].get("active")
    if selected_string:
        if ((listbox_nr==0 or listbox_nr==1) and window_side[0]==False)or((listbox_nr==2 or listbox_nr==3) and window_side[1]==False):
        # Remove the string from the list and update the listbox
            list_arr_NPL[list_nr].remove(selected_string)
            list_arr_NPL[list_nr] = sorted(list_arr_NPL[list_nr])  # Sort the list
            listbox_arr[listbox_nr].delete(0, "end")  # Clear the listbox
            for item in list_arr_NPL[list_nr]:  # Insert the sorted items into the listbox
                listbox_arr[listbox_nr].insert("end", item)

        if ((listbox_nr==0 or listbox_nr==1) and window_side[0]==True)or((listbox_nr==2 or listbox_nr==3) and window_side[1]==True):
                # Remove the string from the list and update the listbox
            list_arr_NPR[list_nr].remove(selected_string)
            list_arr_NPR[list_nr] = sorted(list_arr_NPR[list_nr])  # Sort the list
            listbox_arr[listbox_nr].delete(0, "end")  # Clear the listbox
            for item in list_arr_NPR[list_nr]:  # Insert the sorted items into the listbox
                listbox_arr[listbox_nr].insert("end", item)
                
# Funktion um wieder ins Menü zu kommen
def zurück_button_func():  
    reset_window()
    menu()

# Funktion um die Variable modell_nr zu schreiben
def write_to_variable_modell_nr(entry):
    global modell_nr
    modell_nr = entry.get()
    
# Funktion um die Variable laenge_schuerzen zu schreiben
def write_to_variable_laenge_schuerze(entry):
    laenge_schuerzen[0] = entry.get()

# Programm für die erstellung von Liner 5 Schürzen
def liner5_program(flag):
    # Widgets vom Fenster entfernen 
    for widget in window.winfo_children():
        widget.destroy()
        
    window.title("Dateneingabe Liner 5")

    # Funktion für der Weiterknopf des Liner 5 Programm
    def weiter_button_func():
        
    # Öffnen Sie den Datei-Explorer und speichern Sie den Pfad der ausgewählten Datei
        filepath = filedialog.askopenfilename(filetypes=[('Excel-Dateien', '*.xlsx')], initialdir = script_dir)
            # Öffnen Sie die ausgewählte Excel-Datei mit openpyxl
        wb = openpyxl.load_workbook(filepath)

            # Machen Sie etwas mit der Excel-Datei (z.B. Werte in einer Zelle ändern)
        ws = wb[wb.sheetnames[0]]

        # Sucht die Anzahl von Einträgen in der Excel liste 
        num_rows = 1
        for row in ws.rows:
            if any(cell.value for cell in row):
                num_rows += 1
                
        # schreibt die Variablen in die letzte Zeile der Exceltabelle
        ws.cell(row=num_rows, column=1).value = modell_nr
        ws.cell(row=num_rows, column=2).value = len(list_trennung_alu_NPL)
        ws.cell(row=num_rows, column=3).value = ",".join(str(x) for x in list_trennung_alu_NPL) #list_trennung_alu_NPL anstatt neue variable um add und delete funkrion nicht zu verändern
        ws.cell(row=num_rows, column=4).value = len(list_trennung_schuerze_NPL)
        ws.cell(row=num_rows, column=5).value = ",".join(str(x) for x in list_trennung_schuerze_NPL) #list_trennung_alu_NPL anstatt neue variable um add und delete funkrion nicht zu verändern
        ws.cell(row=num_rows, column=6).value = laenge_schuerzen[0]
        current_date = date.today()
        formatted_date = current_date.strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=num_rows, column=7).value = str(formatted_date)#list_trennung_alu_NPL anstatt neue variable um add und delete funkrion nicht zu verändern

        # speichert die Excel datei und gibt eine Info zur Bestätigung
        wb.save(filepath)
        Antispritzschutz_Script_Liner_5.upload_file(filepath)
        reset_window()
        menu()
        tk.messagebox.showinfo("Confirm", "Daten wurden gespeichert und Zeichnung wurde erstellt!")

    # erstellung der Labels, Entrys, Listboxen und Buttons und die zordnung der Funktionen an die Widgets
    abstand_label = tk.Label(window)
    abstand_label.config(height=1)
    
    modell_nr_label=tk.Label(window, text="Modellnummer")
    modell_nr_entry=tk.Entry(window)
    modell_nr_entry.bind("<FocusOut>", lambda e: write_to_variable_modell_nr(modell_nr_entry))

    länge_label=tk.Label(window, text="Länge Schürzen")
    länge_entry=tk.Entry(window)
    länge_entry.bind("<FocusOut>", lambda e: write_to_variable_laenge_schuerze(länge_entry))

    abstand1_label=tk.Label(window)
    abstand1_label.config(height=4)

    halterungen_label = tk.Label(window, text="Halterungen")
    halterungen_entry = tk.Entry(window)
    halterungen_entry.bind("<Return>", lambda e: add_string(halterungen_entry,0,0))
    add_button = tk.Button(window, text="+", command=lambda: add_string(halterungen_entry,0,0))
    delete_button = tk.Button(window, text="-", command=lambda: delete_string(0,0))
    listbox = tk.Listbox(window, width=25, height=5)

    trennungen_label = tk.Label(window, text="Trennungen")
    trennungen_entry = tk.Entry(window)
    trennungen_entry.bind("<Return>", lambda e: add_string(trennungen_entry,1,1))
    add1_button = tk.Button(window, text="+", command=lambda: add_string(trennungen_entry,1,1))
    delete1_button = tk.Button(window, text="-", command=lambda: delete_string(1,1))
    listbox1 = tk.Listbox(window, width=25, height=5)

    weiter_button = tk.Button(window, text="Weiter", command=weiter_button_func)
    zurück_button = tk.Button(window, text="Zurück", command=zurück_button_func)


    # erstellung des Arrays für die Listboxen
    global listbox_arr
    listbox_arr = [listbox, listbox1]

    # Eintragung der Werte eines ausgewählten Chassis in das Liner 5 Programm
    if flag:
        modell_nr_entry.insert("end", modell_nr)

        if laenge_schuerzen[0]:
            länge_entry.insert("end", laenge_schuerzen[0])

        global list_arr_NPL

        list_trennung_alu_NPL=[float(i) for i in flag[0] if i!="None"]
        list_trennung_schuerze_NPL=[float(i) for i in flag[1] if i!="None"]

        list_arr_NPL= [list_trennung_alu_NPL, list_trennung_schuerze_NPL]   

        for item in list_arr_NPL[0]:  
            if item!="None":
                listbox_arr[0].insert("end", item)

        for item in list_arr_NPL[1]:  
            if item!="None":
                listbox_arr[1].insert("end", item)

    # Ausgabe der Widgets auf dem Fenster
    abstand_label.grid(row=0, column=3)
    
    logo(2,1)

    abstand1_label.grid(column=3,row=3)

    modell_nr_label.grid(column=0,row=6, sticky='nsew', padx=5,pady=5)
    modell_nr_entry.grid(column=1,row=6,columnspan=3, sticky='nsew', padx=5,pady=5)

    länge_label.grid(column=4,row=6, sticky='nsew', padx=5,pady=5)
    länge_entry.grid(column=5,row=6,columnspan=3, sticky='nsew', padx=5,pady=5)

    halterungen_label.grid(column=0, row=8,sticky='nsew', padx=5, pady=5)
    halterungen_entry.grid(column=1, row=8,columnspan=3,sticky='nsew', padx=5, pady=5)
    add_button.grid(row=9, column=0,sticky='nsew', padx=5, pady=5)
    delete_button.grid(row=10, column=0,sticky='nsew', padx=5, pady=5)
    listbox.grid(row=9, column=1, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    trennungen_label.grid(column=4, row=8,sticky='nsew', padx=5, pady=5)
    trennungen_entry.grid(column=5, row=8,columnspan=3,sticky='nsew', padx=5, pady=5)
    add1_button.grid(row=9, column=4,sticky='nsew', padx=5, pady=5)
    delete1_button.grid(row=10, column=4,sticky='nsew', padx=5, pady=5)
    listbox1.grid(row=9, column=5, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    weiter_button.grid(row=14, column=4,columnspan=4, sticky='nsew', padx=5, pady=5)
    zurück_button.grid(row=14, column=0,columnspan=4, sticky='nsew', padx=5, pady=5)
    file_menu()




# Funktion für das Liner 4 Programm
def liner4_program(flag):
    # alte Widgets vom Fenster entfernen
    for widget in window.winfo_children():
        widget.destroy()
        
    window.title("Dateneingabe Liner 4")

    # Funktion zur weitergabe der Werte an die Exceltabelle und das Liner 4 Script
    def weiter_button_func():

        # Öffnen Sie den Datei-Explorer und speichern Sie den Pfad der ausgewählten Datei
        filepath = filedialog.askopenfilename(filetypes=[('Excel-Dateien', '*.xlsx')], initialdir = script_dir)
        
        # Öffnen Sie die ausgewählte Excel-Datei mit openpyxl
        wb = openpyxl.load_workbook(filepath)

        # Machen Sie etwas mit der Excel-Datei (z.B. Werte in einer Zelle ändern)
        ws = wb[wb.sheetnames[0]]

        # Anzahl der Zeilen der Excel Tabelle zählen
        num_rows = 1
        for row in ws.rows:
            if any(cell.value for cell in row):
                num_rows += 1

        # Variablen in die Tabelle schreiben
        ws.cell(row=num_rows, column=1).value = modell_nr
        ws.cell(row=num_rows, column=2).value = laenge_schuerzen[0]
        ws.cell(row=num_rows, column=3).value = laenge_schuerzen[1]
        ws.cell(row=num_rows, column=4).value = len(list_planspanner_NPL)
        ws.cell(row=num_rows, column=5).value = ",".join(str(x) for x in list_planspanner_NPL)
        ws.cell(row=num_rows, column=6).value = len(list_planspanner_NPR)
        ws.cell(row=num_rows, column=7).value = ",".join(str(x) for x in list_planspanner_NPR)
        ws.cell(row=num_rows, column=8).value = len(list_rungen_NPL)
        ws.cell(row=num_rows, column=9).value = ",".join(str(x) for x in list_rungen_NPL)
        ws.cell(row=num_rows, column=10).value = len(list_rungen_NPR)
        ws.cell(row=num_rows, column=11).value = ",".join(str(x) for x in list_rungen_NPR)
        ws.cell(row=num_rows, column=12).value = abstand_alu_schuerze[0]
        ws.cell(row=num_rows, column=13).value = abstand_alu_schuerze[1]
        ws.cell(row=num_rows, column=14).value = len(list_trennung_schuerze_NPL)
        ws.cell(row=num_rows, column=15).value = ",".join(str(x) for x in list_trennung_schuerze_NPL)
        ws.cell(row=num_rows, column=16).value = len(list_trennung_schuerze_NPR)
        ws.cell(row=num_rows, column=17).value = ",".join(str(x) for x in list_trennung_schuerze_NPR)
        ws.cell(row=num_rows, column=18).value = len(list_trennung_alu_NPL)
        ws.cell(row=num_rows, column=19).value = ",".join(str(x) for x in list_trennung_alu_NPL)
        ws.cell(row=num_rows, column=20).value = len(list_trennung_alu_NPR)
        ws.cell(row=num_rows, column=21).value = ",".join(str(x) for x in list_trennung_alu_NPR)
        ws.cell(row=num_rows, column=22).value = float(angewinkelt_links.get())
        ws.cell(row=num_rows, column=23).value = float(angewinkelt_rechts.get())
        current_date = date.today()
        formatted_date = current_date.strftime('%Y-%m-%d %H:%M:%S')
        ws.cell(row=num_rows, column=24).value = str(formatted_date)


        # Speichern Sie das Dokument
        wb.save(filepath)
        Antispritzschutz_Script_Liner_4.upload_file(filepath)
        reset_window()
        menu()
        tk.messagebox.showinfo("Confirm", "Daten wurden gespeichert und Zeichnung wurde erstellt!")

    # Funktion zur speicherung der Variabeln falls der Button für den Nullpunkt links aktiv ist
    def nullpunkt_links(x):
        button_arr[x+1].configure(bg="SystemButtonFace")
        button_arr[x].configure(bg="grey") 
        if x==0:
            window_side[0]=False

            abstand_alu_schuerze_entry.delete(0,"end")
            abstand_alu_schuerze_entry.insert(0, abstand_alu_schuerze[0])

            listbox_arr[0].delete(0, "end")  # Clear the listbox
            listbox_arr[1].delete(0, "end")  # Clear the listbox

            for item in list_arr_NPL[0]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[0].insert("end", item)
            for item in list_arr_NPL[1]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[1].insert("end", item)
        else:
            window_side[1]=False

            laenge_schuerze_entry.delete(0,"end")
            laenge_schuerze_entry.insert(0, laenge_schuerzen[0])

            listbox_arr[2].delete(0, "end")  # Clear the listbox
            listbox_arr[3].delete(0, "end")  # Clear the listbox

            for item in list_arr_NPL[2]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[2].insert("end", item)
            for item in list_arr_NPL[3]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[3].insert("end", item)    

    # Funktion zur speicherung der Variabeln falls der Button für den Nullpunkt rechts aktiv ist
    def nullpunkt_rechts(x):
        button_arr[x].configure(bg="grey")
        button_arr[x-1].configure(bg="SystemButtonFace")
        if x==1:
            window_side[0]=True

            abstand_alu_schuerze_entry.delete(0,"end")
            abstand_alu_schuerze_entry.insert(0, abstand_alu_schuerze[1])
        
            listbox_arr[0].delete(0, "end")  # Clear the listbox
            listbox_arr[1].delete(0, "end")  # Clear the listbox

            for item in list_arr_NPR[0]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[0].insert("end", item)
            for item in list_arr_NPR[1]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[1].insert("end", item)
        else:
            window_side[1]=True   
            
            laenge_schuerze_entry.delete(0,"end")
            laenge_schuerze_entry.insert(0, laenge_schuerzen[1])

            listbox_arr[2].delete(0, "end")  # Clear the listbox
            listbox_arr[3].delete(0, "end")  # Clear the listbox

            for item in list_arr_NPR[2]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[2].insert("end", item)
            for item in list_arr_NPR[3]:  # Insert the sorted items into the listbox
                if item!="None":
                    listbox_arr[3].insert("end", item) 
    
    # Funktion um die Variable abstand_alu_schuerze zu schreiben
    def write_to_variable_abstand_alu_schuerze(event):
        value = abstand_alu_schuerze_entry.get()

        if window_side[0]==True:
            abstand_alu_schuerze[1]=value
        else:
            abstand_alu_schuerze[0]=value
            
    # Funktion um die Variable laenge_schuerzen zu schreiben
    def write_to_variable_laenge_schuerze(event):
        value = laenge_schuerze_entry.get()
    
        if window_side[1]==True:
            laenge_schuerzen[1]=value
        else:
            laenge_schuerzen[0]=value

    # widgets erstellen und Funktionen anbinden
    aluleiste_label = tk.Label(window, text="Aluleiste")
    schuerze_label = tk.Label(window, text="Kunststoffschürze")

    nullpunkt_links0_button = tk.Button(window, text="Nullpunkt Links", command=lambda:nullpunkt_links(0), bg="grey")
    nullpunkt_rechts0_button = tk.Button(window, text="Nullpunkt Rechts", command=lambda:nullpunkt_rechts(1))
    nullpunkt_links1_button = tk.Button(window, text="Nullpunkt Links", command=lambda:nullpunkt_links(2), bg="grey")
    nullpunkt_rechts1_button = tk.Button(window, text="Nullpunkt Rechts", command=lambda:nullpunkt_rechts(3))
    
    modell_nr_label=tk.Label(window, text="Modellnummer")
    modell_nr_entry=tk.Entry(window)
    modell_nr_entry.bind("<FocusOut>", lambda e: write_to_variable_modell_nr(modell_nr_entry))

    angewinkelt_links_check = tk.Checkbutton(window, text="Winkel Links", variable=angewinkelt_links)
    angewinkelt_rechts_check = tk.Checkbutton(window, text="Winkel Rechts", variable=angewinkelt_rechts)

    abstand_alu_schuerze_label = tk.Label(window, text="<--> NP Schürze")
    abstand_alu_schuerze_entry = tk.Entry(window)
    abstand_alu_schuerze_entry.bind("<FocusOut>", write_to_variable_abstand_alu_schuerze)

    laenge_schuerze_label = tk.Label(window, text="Länge")
    laenge_schuerze_entry = tk.Entry(window)
    laenge_schuerze_entry.bind("<FocusOut>", write_to_variable_laenge_schuerze)

    trennung_alu_label = tk.Label(window, text="Trennungen")
    trennung_alu_entry = tk.Entry(window)
    trennung_alu_entry.bind("<Return>", lambda e: add_string(trennung_alu_entry,0,0))
    add_button = tk.Button(window, text="+", command=lambda: add_string(trennung_alu_entry,0,0))
    delete_button = tk.Button(window, text="-", command=lambda: delete_string(0,0))
    listbox = tk.Listbox(window, width=25, height=5)

    rungen_label = tk.Label(window, text="Rungen")
    rungen_entry = tk.Entry(window)
    rungen_entry.bind("<Return>", lambda e: add_string(rungen_entry,1,1))
    add1_button = tk.Button(window, text="+", command=lambda: add_string(rungen_entry,1,1))
    delete1_button = tk.Button(window, text="-", command=lambda: delete_string(1,1))
    listbox1 = tk.Listbox(window, width=25, height=5)

    trennung1_label = tk.Label(window, text="Trennungen")
    trennung1_entry = tk.Entry(window)
    trennung1_entry.bind("<Return>", lambda e: add_string(trennung1_entry,2,2))
    add2_button = tk.Button(window, text="+", command=lambda: add_string(trennung1_entry,2,2))
    delete2_button = tk.Button(window, text="-", command=lambda: delete_string(2,2))
    listbox2 = tk.Listbox(window, width=25, height=5)

    planspanner_label = tk.Label(window, text="Planspanner")
    planspanner_entry = tk.Entry(window)
    planspanner_entry.bind("<Return>", lambda e: add_string(planspanner_entry,3,3))
    add3_button = tk.Button(window, text="+", command=lambda: add_string(planspanner_entry,3,3))
    delete3_button = tk.Button(window, text="-", command=lambda: delete_string(3,3))
    listbox3 = tk.Listbox(window, width=25, height=5)

    weiter_button = tk.Button(window, text="Weiter", command=weiter_button_func)
    zurück_button = tk.Button(window, text="Zurück", command=zurück_button_func)

    # widgets in arrays ablegen
    global listbox_arr
    listbox_arr = [listbox,listbox1,listbox2,listbox3]
    button_arr= [nullpunkt_links0_button,nullpunkt_rechts0_button,nullpunkt_links1_button,nullpunkt_rechts1_button]
    
    # Eintragung der Werte eines ausgewählten Chassis in das Liner 4 Programm
    if flag:
        modell_nr_entry.insert("end", modell_nr)

        if laenge_schuerzen[0]:
            laenge_schuerze_entry.insert("end", laenge_schuerzen[0])

        if abstand_alu_schuerze[0]:
            abstand_alu_schuerze_entry.insert("end", abstand_alu_schuerze[0])

        global list_arr_NPL
        global list_arr_NPR

        list_trennung_alu_NPL=[float(i) for i in flag[0] if i!="None"]
        list_rungen_NPL=[float(i) for i in flag[1] if i!="None"]
        list_trennung_schuerze_NPL=[float(i) for i in flag[2] if i!="None"]
        list_planspanner_NPL=[float(i) for i in flag[3] if i!="None"]
        list_trennung_alu_NPR=[float(i) for i in flag[4] if i!="None"]
        list_rungen_NPR=[float(i) for i in flag[5] if i!="None"]
        list_trennung_schuerze_NPR=[float(i) for i in flag[6] if i!="None"]
        list_planspanner_NPR=[float(i) for i in flag[7] if i!="None"]

        list_arr_NPL= [list_trennung_alu_NPL, list_rungen_NPL, list_trennung_schuerze_NPL, list_planspanner_NPL]   
        list_arr_NPR= [list_trennung_alu_NPR, list_rungen_NPR, list_trennung_schuerze_NPR, list_planspanner_NPR]   

        for item in list_arr_NPL[0]:  # Insert the sorted items into the listbox
            if item!="None":
                listbox_arr[0].insert("end", item)

        for item in list_arr_NPL[1]:  # Insert the sorted items into the listbox
            if item!="None":
                listbox_arr[1].insert("end", item)

        for item in list_arr_NPL[2]:  # Insert the sorted items into the listbox
            if item!="None":
                listbox_arr[2].insert("end", item)

        for item in list_arr_NPL[3]:  # Insert the sorted items into the listbox
            if item!="None":
                listbox_arr[3].insert("end", item) 

    #widgets anzeigen
    abstand_label = tk.Label(window)
    abstand_label.config(height=1)
    abstand_label.grid(row=0, column=3)

    logo(2,1)

    aluleiste_label.grid(column=0, row=4, columnspan=4,sticky='nsew', padx=5, pady=5)
    schuerze_label.grid(column=4, row=4, columnspan=4,sticky='nsew', padx=5, pady=5)

    nullpunkt_links0_button.grid(row=5, column=0, columnspan=2,sticky='nsew', padx=5, pady=5)
    nullpunkt_rechts0_button.grid(row=5, column=2, columnspan=2,sticky='nsew', padx=5, pady=5)
    nullpunkt_links1_button.grid(row=5, column=4, columnspan=2,sticky='nsew', padx=5, pady=5)
    nullpunkt_rechts1_button.grid(row=5, column=6, columnspan=2,sticky='nsew', padx=5, pady=5)

    modell_nr_label.grid(column=0,row=6, sticky='nsew', padx=5,pady=5)
    modell_nr_entry.grid(column=1,row=6,columnspan=3, sticky='nsew', padx=5,pady=5)

    angewinkelt_links_check.grid(column=4, row=6,columnspan=2,sticky='nsew', padx=5, pady=5)
    angewinkelt_rechts_check.grid(column=6, row=6,columnspan=2,sticky='nsew', padx=5, pady=5)

    abstand_alu_schuerze_label.grid(column=0, row=7,sticky='nsew', padx=5, pady=5)
    abstand_alu_schuerze_entry.grid(column=1, row=7,columnspan=3,sticky='nsew', padx=5, pady=5)

    laenge_schuerze_label.grid(column=4, row=7,sticky='nsew', padx=5, pady=5)
    laenge_schuerze_entry.grid(column=5, row=7,columnspan=3,sticky='nsew', padx=5, pady=5)

    trennung_alu_label.grid(column=0, row=8,sticky='nsew', padx=5, pady=5)
    trennung_alu_entry.grid(column=1, row=8,columnspan=3,sticky='nsew', padx=5, pady=5)
    add_button.grid(row=9, column=0,sticky='nsew', padx=5, pady=5)
    delete_button.grid(row=10, column=0,sticky='nsew', padx=5, pady=5)
    listbox.grid(row=9, column=1, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    rungen_label.grid(column=0, row=11,sticky='nsew', padx=5, pady=5)
    rungen_entry.grid(column=1, row=11,columnspan=3,sticky='nsew', padx=5, pady=5)
    add1_button.grid(row=12, column=0,sticky='nsew', padx=5, pady=5)
    delete1_button.grid(row=13, column=0,sticky='nsew', padx=5, pady=5)
    listbox1.grid(row=12, column=1, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    trennung1_label.grid(column=4, row=8,sticky='nsew', padx=5, pady=5)
    trennung1_entry.grid(column=5, row=8,columnspan=3,sticky='nsew', padx=5, pady=5)
    add2_button.grid(row=9, column=4,sticky='nsew', padx=5, pady=5)
    delete2_button.grid(row=10, column=4,sticky='nsew', padx=5, pady=5)
    listbox2.grid(row=9, column=5, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    planspanner_label.grid(column=4, row=11,sticky='nsew', padx=5, pady=5)
    planspanner_entry.grid(column=5, row=11,columnspan=3,sticky='nsew', padx=5, pady=5)
    add3_button.grid(row=12, column=4,sticky='nsew', padx=5, pady=5)
    delete3_button.grid(row=13, column=4,sticky='nsew', padx=5, pady=5)
    listbox3.grid(row=12, column=5, rowspan=2, columnspan=3,sticky='nsew', padx=5, pady=5)

    weiter_button.grid(row=14, column=4,columnspan=4, sticky='nsew', padx=5, pady=5)
    zurück_button.grid(row=14, column=0,columnspan=4, sticky='nsew', padx=5, pady=5)
    file_menu()
# Funktion zur Anzeige des Krone Logos
def logo(x,y):
    img_path = os.path.join(os.path.dirname(__file__), 'Bilder', 'Krone_Logo-removebg-preview.png')
    logo = Image.open(img_path)
    logo = logo.resize((int(984/x),int(253/x)), resample=Image.Resampling.LANCZOS)
    photo = ImageTk.PhotoImage(logo)
    logo_label= tk.Label(image=photo)
    logo_label.image=photo
    logo_label.grid(row=y, rowspan=2, column=1, columnspan=6, sticky='nsew')

# Funktion zum beenden des Programms
def beenden():
    window.destroy()

# Funktion zur Anzeige des Menüs
def menu():
# Create a start button
    reset_window()

    window.title("Antispritzschutz Schablonen Generator")

    menu_label = tk.Label(window)
    menu_label.config(height=4)
    menu_label.grid(row=0, column=3)

    logo(2,2)

    menu_label = tk.Label(window, text= "Programm zur automatischen Erstellung von Schablonen für die Antispritzschutzschürzen")
    menu_label.grid(row=4, column=1, columnspan=6, sticky='nsew')
    menu_label.config(font=("Helvetica", 10, "bold"),height=5)

    abstand_label = tk.Label(window)
    abstand_label.config(height=4)
    abstand_label.grid(row=5, column=3)

    liner4_button = tk.Button(window, text="Programm Liner 4", command=lambda:liner4_program(0))
    liner4_button.config(height=2)
    liner4_button.grid(row=10, rowspan=3, column=3, columnspan=2, sticky='nsew', padx=5, pady=5)
    liner5_button = tk.Button(window, text="Programm Liner 5", command=lambda:liner5_program(0))
    liner5_button.config(height=2)
    liner5_button.grid(row=13, rowspan= 3, column=3, columnspan=2, sticky='nsew', padx=5, pady=5)
    exit_button = tk.Button(window, text="Beenden", command=lambda:beenden())
    exit_button.config(height=2)
    exit_button.grid(row=16, rowspan= 3, column=3, columnspan=2, sticky='nsew', padx=5, pady=5)
    file_menu()

# Funktion für das File Menü
def file_menu():
    
    menubar = tk.Menu(window)
    window.config(menu=menubar)

# Add a File menu
    file_menu = tk.Menu(menubar)
    menubar.add_cascade(label='Menü', menu=file_menu)

# Add an Open option to the File menu
    file_menu.add_command(label='Datei öffnen', command=open_file)
    file_menu.add_command(label='Liner 4', command=lambda:liner4_program(0))
    file_menu.add_command(label='Liner 5', command=liner5_program)
    file_menu.add_command(label='Beenden', command=beenden)


# Funktion zur Weitergabe des ausgesuchten Chassis
def open_file():
    
    # Exceltabelle auswählen und ein neues Fenster für das Auswählen der Chassis
    filepath = filedialog.askopenfile(filetypes=[('Excel-Dateien', '*.xlsx')], initialdir = script_dir)
    window2=tk.Tk()
    window_width=400
    window_height=280

    window2.title("Dokument öffnen")

    # Fenster in der Mitte des Bildschirms ausrichten und anzahl der Spalten auswählen
    screen_width = window2.winfo_screenwidth()
    screen_height = window2.winfo_screenheight()
    x_coord = (screen_width/2) - (window_width/2)
    y_coord = (screen_height/2) - (window_height/2) -50
    window2.geometry("+{}+{}".format(int(x_coord), int(y_coord)))

    window2.configure(highlightthickness=3, highlightbackground='black')
    for i in range(1,5):
        window2.columnconfigure(i, weight=1)
    window2.geometry(str(window_width) + "x" + str(window_height))

    # Widgets erstellen und Funtionen anfügen
    listbox = tk.Listbox(window2, width=25, height=7)
    listbox2 = tk.Listbox(window2, width=25, height=7)
    listbox.bind('<Button-1>',lambda e: anzeige(listbox))
    listbox2.bind('<Button-1>',lambda e: anzeige(listbox2))

    # Funktion um das ausgewählte Chassi in der Suchzeile anzuzeigen
    def anzeige(listboxnr):
        selected_string = listboxnr.get("active").split(' ')
        query_entry.delete(0,END)
        query_entry.insert(0,str(selected_string[0]))

    # Excel Tabelle öffnen und Werte vom ausgewählten Chassi kopieren
    wb = openpyxl.load_workbook(filepath.name)
    ws = wb[wb.sheetnames[0]]
    row_count = 0
    row_count2 = 1
    
    name=filepath.name.split("/")
    if name[-1]== "Antispritzschutz_Daten_Liner_4.xlsx":
        date=23
    elif name[-1]== "Antispritzschutz_Daten_Liner_5.xlsx":
        date=6
        
    # kopierte Werte in die Listboxen einfügen
    for rows in ws.rows:
        #print(rows)
        if row_count == 0:  # Skip the first row
            row_count += 1
            continue
        row_count2+=1

        date_time_obj = datetime.strptime(str(rows[date].value), '%Y-%m-%d %H:%M:%S')
        date_obj = date_time_obj.date()
        
        if row_count2<=len(list(ws.rows))/2+1:
            string = str(rows[0].value).ljust(20) + str(date_obj).rjust(20)
            listbox.insert(END, string)

        else:
            # string = "{:<20}{:>20}".format(str(rows[0].value), str(date_obj))
            string = str(rows[0].value).ljust(20) + str(date_obj).rjust(20)

            listbox2.insert(END, string)

    # Widgets anzeigen lassen
    suche_label = tk.Label(window2, text="Suche:")
    suche_label.config(height=2)
    suche_label.grid(row=0, column=1,sticky='sw', padx=5, pady=5)
    
    query_entry = tk.Entry(window2)
    query_entry.grid(column=1, row=1,columnspan=4,sticky='nsew', padx=5, pady=5)

    modell_label = tk.Label(window2, text="Wählen sie ein Modell aus:")
    modell_label.config(height=1)
    modell_label.grid(row=2, column=1,sticky='sw', padx=5, pady=5)

    listbox.grid(row=3, column=1, rowspan=4, columnspan=2,sticky='nsew', padx=5, pady=5)
    
    listbox2.grid(row=3, column=3, rowspan=4, columnspan=2,sticky='nsew', padx=5, pady=5)
    zurück_button = tk.Button(window2, text="Zurück", command=lambda:window2.destroy())
    zurück_button.grid(row=7, column=1, columnspan=2, sticky='nsew', padx=5, pady=5)
    
    if date==23: 
        weiter_l4_button = tk.Button(window2, text="Weiter", command=lambda:weiter_menu_liner_4(ws,query_entry,window,window2))
        weiter_l4_button.grid(row=7, column=3, columnspan=2, sticky='nsew', padx=5, pady=5)
       
    else:
        weiter_l5_button = tk.Button(window2, text="Weiter", command=lambda:weiter_menu_liner_5(ws,query_entry,window,window2))
        weiter_l5_button.grid(row=7, column=3, columnspan=2, sticky='nsew', padx=5, pady=5)
        

# Funktion zur Weitergabe der Daten an das nächste Script
def weiter_menu_liner_4(ws,query,window,window2):
    # Get the string from the entry widget
    string=str(query.get())
    row_count = 0
    row=[]
    for rows in ws.rows:
    #print(rows)
        if row_count == 0:  # Skip the first row
            row_count += 1
            continue
        if rows[0].value==string:
            for cell in rows:
                row.append(cell.value)
            break

    global modell_nr
    modell_nr = row[0]
    if row[1]!=None:
        laenge_schuerzen[0] = float(row[1])
    if row[2]!=None:
        laenge_schuerzen[1] = float(row[2])
    if row[11]!=None:
        abstand_alu_schuerze[0] = float(row[11])
    if row[12]!=None:
        abstand_alu_schuerze[1] = float(row[12])

    list_planspanner_NPL, list_planspanner_NPR, list_rungen_NPL, list_rungen_NPR, list_trennung_schuerze_NPL, list_trennung_schuerze_NPR, list_trennung_alu_NPL, list_trennung_alu_NPR = [
    str(rows[i].value).split(',') if str(rows[i].value) and not str(rows[i].value).isspace() else []
    for i in [4, 6, 8, 10, 14, 16, 18, 20]]
    
    list_arr= [list_trennung_alu_NPL, list_rungen_NPL, list_trennung_schuerze_NPL, list_planspanner_NPL, list_trennung_alu_NPR, list_rungen_NPR, list_trennung_schuerze_NPR, list_planspanner_NPR]   

    for widget in window.winfo_children():
        widget.destroy()
    window2.destroy()
    liner4_program(list_arr)
    
# Funktion zur Weitergabe der Daten an das nächste Script
def weiter_menu_liner_5(ws,query,window,window2):
    # Get the string from the entry widget
    string=str(query.get())
    row_count = 0
    row=[]
    for rows in ws.rows:
    #print(rows)
        if row_count == 0:  # Skip the first row
            row_count += 1
            continue
        if rows[0].value==string:
            for cell in rows:
                row.append(cell.value)
            break

    global modell_nr
    modell_nr = row[0]
    if row[5]!=None:
        laenge_schuerzen[0] = float(row[5])

    list_trennung_alu_NPL, list_trennung_schuerze_NPL = [
    str(rows[i].value).split(',') if str(rows[i].value) and not str(rows[i].value).isspace() else []
    for i in [2, 4]]
    
    list_arr= [list_trennung_alu_NPL, list_trennung_schuerze_NPL]   

    for widget in window.winfo_children():
        widget.destroy()
    window2.destroy()
    liner5_program(list_arr)

def setup():

    global script_dir
    
    if os.path.splitext(sys.argv[0])[1] == '.exe':
        script_dir = os.path.dirname(os.path.realpath(sys.executable))
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    if not os.path.exists("Liner 4 Schürzen"):
        os.makedirs("Liner 4 Schürzen")
    if not os.path.exists("Liner 5 Schürzen"):
        os.makedirs("Liner 5 Schürzen")

    if not os.path.exists("Antispritzschutz_Daten_Liner_4.xlsx"):
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = "Antispritzschutz_Daten_Liner_4"
        ws1.append(["Modell Nr.", "Länge Schürze NPL", "Länge Schürze NPR", "Anzahl Planspanner NPL", "Pos. Planspanner NPL", "Anzahl Planspanner NPR", "Pos. Planspanner NPR", "Anzahl Rungen NPL", "Pos. Rungen NPL", "Anzahl Rungen NPR", "Pos. Rungen NPR", "Abstand Alu/Schürze NPL", "Abstand Alu/Schürze NPR", "Anzahl Trennungen Schürze NPL", "Pos. Trennungen Schürze NPL", "Anzahl Trennungen Schürze NPR", "Pos. Trennungen Schürze NPR", "Anzahl Trennungen Alu NPL", "Pos. Trennungen Alu NPL", "Anzahl Trennungen Alu NPR", "Pos. Trennungen Alu NPR", "Winkel Links", "Winkel Rechts","Erstellungsdatum"])
        wb1.save(filename = "Antispritzschutz_Daten_Liner_4.xlsx")

    if not os.path.exists("Antispritzschutz_Daten_Liner_5.xlsx"):
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "Antispritzschutz_Daten_Liner_5"
        ws2.append(["Modell Nr.", "Anzahl Halterungen", "Pos. Halterungen", "Anzahl Trennungen", "Pos. Trennungen", "Länge Schürze","Ertstellungsdatum"])
        wb2.save(filename = "Antispritzschutz_Daten_Liner_5.xlsx")


if __name__ == "__main__":
    setup()
    menu()
    

# Run the main loop
window.mainloop()

